"""FastAPI service to translate Excel header rows from Chinese to English."""

import io
from tempfile import NamedTemporaryFile

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse
import openpyxl
from deep_translator import GoogleTranslator

app = FastAPI(title="Excel Header Translator", description="Translate Excel headers from Chinese to English")


@app.get("/", response_class=HTMLResponse)
async def home():
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Excel Header Translator</title>
        <style>
            * { box-sizing: border-box; margin: 0; padding: 0; }
            body { font-family: -apple-system, sans-serif; background: #f5f5f5; display: flex;
                   justify-content: center; align-items: center; min-height: 100vh; }
            .card { background: #fff; border-radius: 12px; padding: 40px; max-width: 480px;
                    width: 90%; box-shadow: 0 4px 24px rgba(0,0,0,0.1); text-align: center; }
            h1 { margin-bottom: 8px; font-size: 24px; }
            p { color: #666; margin-bottom: 24px; }
            .upload-area { border: 2px dashed #ccc; border-radius: 8px; padding: 40px 20px;
                           cursor: pointer; transition: border-color 0.2s; margin-bottom: 20px; }
            .upload-area:hover { border-color: #4f46e5; }
            .upload-area.dragover { border-color: #4f46e5; background: #f0f0ff; }
            input[type=file] { display: none; }
            button { background: #4f46e5; color: #fff; border: none; padding: 12px 32px;
                     border-radius: 8px; font-size: 16px; cursor: pointer; }
            button:disabled { background: #aaa; cursor: not-allowed; }
            .status { margin-top: 16px; color: #333; }
        </style>
    </head>
    <body>
        <div class="card">
            <h1>Excel Header Translator</h1>
            <p>Upload an .xlsx file to translate Chinese headers to English</p>
            <form id="form">
                <div class="upload-area" id="drop">
                    <p id="label">Click or drag & drop an .xlsx file here</p>
                    <input type="file" id="file" accept=".xlsx">
                </div>
                <button type="submit" id="btn" disabled>Translate & Download</button>
            </form>
            <div class="status" id="status"></div>
        </div>
        <script>
            const drop = document.getElementById('drop');
            const fileInput = document.getElementById('file');
            const btn = document.getElementById('btn');
            const label = document.getElementById('label');
            const status = document.getElementById('status');

            drop.addEventListener('click', () => fileInput.click());
            drop.addEventListener('dragover', e => { e.preventDefault(); drop.classList.add('dragover'); });
            drop.addEventListener('dragleave', () => drop.classList.remove('dragover'));
            drop.addEventListener('drop', e => {
                e.preventDefault(); drop.classList.remove('dragover');
                fileInput.files = e.dataTransfer.files;
                fileInput.dispatchEvent(new Event('change'));
            });
            fileInput.addEventListener('change', () => {
                if (fileInput.files.length) {
                    label.textContent = fileInput.files[0].name;
                    btn.disabled = false;
                }
            });
            document.getElementById('form').addEventListener('submit', async e => {
                e.preventDefault();
                btn.disabled = true;
                status.textContent = 'Translating...';
                const fd = new FormData();
                fd.append('file', fileInput.files[0]);
                try {
                    const res = await fetch('/translate', { method: 'POST', body: fd });
                    if (!res.ok) { const err = await res.json(); throw new Error(err.detail); }
                    const blob = await res.blob();
                    const a = document.createElement('a');
                    a.href = URL.createObjectURL(blob);
                    a.download = fileInput.files[0].name.replace('.xlsx', '_translated.xlsx');
                    a.click();
                    status.textContent = 'Done! Check your downloads.';
                } catch (err) { status.textContent = 'Error: ' + err.message; }
                btn.disabled = false;
            });
        </script>
    </body>
    </html>
    """


@app.post("/translate")
async def translate(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are supported")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Collect unique text values from header rows (row 1 and row 2)
    header_rows = [1, 2]
    unique_texts = {}
    for row_num in header_rows:
        for cell in ws[row_num]:
            if cell.value and isinstance(cell.value, str):
                unique_texts[cell.value] = None

    if not unique_texts:
        raise HTTPException(400, "No text headers found in rows 1-2")

    translator = GoogleTranslator(source="zh-CN", target="en")
    texts = list(unique_texts.keys())
    translated = translator.translate_batch(texts)
    translation_map = dict(zip(texts, translated))

    for row_num in header_rows:
        for cell in ws[row_num]:
            if cell.value in translation_map:
                cell.value = translation_map[cell.value]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={file.filename.replace('.xlsx', '_translated.xlsx')}"},
    )
